# -*- coding: utf-8 -*-
"""
Raporlar Modülü
Rapor oluşturma ve Excel'e aktarım
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import date, timedelta
import traceback
from tkcalendar import DateEntry
from database import Database
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class Raporlar:
    """Raporlar modülü"""
    
    def __init__(self, parent):
        self.parent = parent
        self.current_report_data = []
        self.create_ui()
    
    def create_ui(self):
        """UI oluştur"""
        # Ana container
        main_frame = tk.Frame(self.parent, bg='#f0f0f0')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Başlık
        header = tk.Label(
            main_frame,
            text="📈 Raporlar",
            font=("Segoe UI", 20, "bold"),
            bg='#f0f0f0',
            fg='#2c3e50'
        )
        header.pack(anchor='w', pady=(0, 20))
        
        # Üst panel - Filtreler
        self.create_filter_panel(main_frame)
        
        # Alt panel - Rapor görüntüleme
        self.create_report_panel(main_frame)
    
    def create_filter_panel(self, parent):
        """Filtre paneli"""
        filter_frame = tk.LabelFrame(
            parent,
            text="Rapor Ayarları",
            font=("Segoe UI", 12, "bold"),
            bg='white',
            fg='#2c3e50',
            relief=tk.FLAT,
            bd=2
        )
        filter_frame.pack(fill=tk.X, pady=(0, 15))
        
        content = tk.Frame(filter_frame, bg='white')
        content.pack(fill=tk.BOTH, padx=15, pady=15)
        
        # Rapor tipi
        tk.Label(
            content,
            text="Rapor Tipi:",
            font=("Segoe UI", 10, "bold"),
            bg='white'
        ).grid(row=0, column=0, sticky='w', pady=(0, 5))
        
        self.report_type_var = tk.StringVar(value='malzeme_cikis')
        report_types = [
            ('malzeme_giris', 'Malzeme Giriş Raporu'),
            ('malzeme_cikis', 'Malzeme Çıkış Raporu'),
            ('urun_bazli', 'Ürün Bazlı Rapor'),
            ('gunluk_ozet', 'Günlük Özet Rapor')
        ]
        
        type_frame = tk.Frame(content, bg='white')
        type_frame.grid(row=1, column=0, columnspan=3, sticky='w', pady=(0, 15))
        
        for value, text in report_types:
            tk.Radiobutton(
                type_frame,
                text=text,
                variable=self.report_type_var,
                value=value,
                font=("Segoe UI", 9),
                bg='white',
                selectcolor='#2196F3'
            ).pack(side=tk.LEFT, padx=(0, 15))
        
        # Tarih aralığı
        tk.Label(
            content,
            text="Başlangıç Tarihi:",
            font=("Segoe UI", 10, "bold"),
            bg='white'
        ).grid(row=2, column=0, sticky='w', pady=(0, 5))
        
        # Varsayılan: Bu ayın ilk günü
        first_day = date.today().replace(day=1)
        
        self.start_date = DateEntry(
            content,
            width=15,
            background='#2196F3',
            foreground='white',
            borderwidth=2,
            date_pattern='dd.mm.yyyy',
            font=("Segoe UI", 10)
        )
        self.start_date.set_date(first_day)
        self.start_date.grid(row=3, column=0, sticky='w', pady=(0, 15))
        
        tk.Label(
            content,
            text="Bitiş Tarihi:",
            font=("Segoe UI", 10, "bold"),
            bg='white'
        ).grid(row=2, column=1, sticky='w', padx=(20, 0), pady=(0, 5))
        
        self.end_date = DateEntry(
            content,
            width=15,
            background='#2196F3',
            foreground='white',
            borderwidth=2,
            date_pattern='dd.mm.yyyy',
            font=("Segoe UI", 10)
        )
        self.end_date.grid(row=3, column=1, sticky='w', padx=(20, 0), pady=(0, 15))
        
        # Butonlar
        btn_frame = tk.Frame(content, bg='white')
        btn_frame.grid(row=3, column=2, sticky='e', padx=(20, 0))
        
        tk.Button(
            btn_frame,
            text="📊 Rapor Oluştur",
            font=("Segoe UI", 10, "bold"),
            bg='#2196F3',
            fg='white',
            cursor='hand2',
            relief=tk.FLAT,
            command=self.generate_report,
            padx=20,
            pady=8
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        tk.Button(
            btn_frame,
            text="📥 Excel'e Aktar",
            font=("Segoe UI", 10),
            bg='#4CAF50',
            fg='white',
            cursor='hand2',
            relief=tk.FLAT,
            command=self.export_to_excel,
            padx=20,
            pady=8
        ).pack(side=tk.LEFT)
        
        content.grid_columnconfigure(2, weight=1)
    
    def create_report_panel(self, parent):
        """Rapor görüntüleme paneli"""
        report_frame = tk.LabelFrame(
            parent,
            text="Rapor Sonuçları",
            font=("Segoe UI", 12, "bold"),
            bg='white',
            fg='#2c3e50',
            relief=tk.FLAT,
            bd=2
        )
        report_frame.pack(fill=tk.BOTH, expand=True)
        
        # Treeview
        tree_frame = tk.Frame(report_frame, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        scrollbar_y = ttk.Scrollbar(tree_frame)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        
        scrollbar_x = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.tree = ttk.Treeview(
            tree_frame,
            show="headings",
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set,
            selectmode="browse"
        )
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_y.config(command=self.tree.yview)
        scrollbar_x.config(command=self.tree.xview)
    
    def generate_report(self):
        """Rapor oluştur"""
        report_type = self.report_type_var.get()
        start = self.start_date.get_date().strftime('%Y-%m-%d')
        end = self.end_date.get_date().strftime('%Y-%m-%d')
        
        try:
            with Database() as db:
                if report_type == 'malzeme_giris':
                    data = db.get_malzeme_giris_raporu(start, end)
                    columns = ["Tarih", "Ürün", "Giriş Miktarı", "Birim", "Fiyat", "Tutar"]
                    self.display_report(data, columns, 
                                      ['tarih', 'cinsi', 'giris_miktari', 'birimi', 'fiyati', 'tutar'])
                
                elif report_type == 'malzeme_cikis':
                    data = db.get_malzeme_cikis_raporu(start, end)
                    columns = ["Tarih", "Öğün", "Ürün", "Çıkış Miktarı", "Birim", "Fiyat", "Tutar", "Kişi Sayısı"]
                    self.display_report(data, columns,
                                      ['tarih', 'ogun', 'cinsi', 'cikis_miktari', 'birimi', 'fiyati', 'tutar', 'mevcut'])
                
                elif report_type == 'urun_bazli':
                    data = db.get_urun_bazli_rapor(start, end)
                    columns = ["Ürün", "Birim", "Toplam Giriş", "Toplam Çıkış", "Toplam Tutar", "İşlem Sayısı"]
                    self.display_report(data, columns,
                                      ['cinsi', 'birimi', 'toplam_giris', 'toplam_cikis', 'toplam_tutar', 'islem_sayisi'])
                
                elif report_type == 'gunluk_ozet':
                    # Günlük özet için özel sorgu
                    data = db.execute_query("""
                        SELECT 
                            tarih,
                            ogun,
                            COUNT(*) as urun_sayisi,
                            SUM(verilen) as toplam_miktar,
                            SUM(tutar) as toplam_tutar,
                            AVG(sahis_tutar) as ortalama_kisi_tutar,
                            SUM(sahis_kalori) as toplam_kalori
                        FROM gunluk_tabela
                        WHERE tarih BETWEEN ? AND ?
                        GROUP BY tarih, ogun
                        ORDER BY tarih, ogun
                    """, (start, end))
                    columns = ["Tarih", "Öğün", "Ürün Sayısı", "Toplam Miktar", "Toplam Tutar", "Ort. Kişi Tutar", "Toplam Kalori"]
                    self.display_report(data, columns,
                                      ['tarih', 'ogun', 'urun_sayisi', 'toplam_miktar', 'toplam_tutar', 'ortalama_kisi_tutar', 'toplam_kalori'])
                
                self.current_report_data = data
                
                if not data:
                    messagebox.showinfo("Bilgi", "Seçili tarih aralığında veri bulunamadı.")
                    
        except Exception as e:
            messagebox.showerror("Hata", f"Rapor oluşturulamadı:\n{e}")
            traceback.print_exc()
    
    def display_report(self, data, columns, field_names):
        """Raporu görüntüle"""
        # Treeview'i temizle
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Kolonları ayarla
        self.tree["columns"] = columns
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120, anchor='center')
        
        # Verileri ekle
        for row in data:
            values = []
            for field in field_names:
                val = row.get(field, '')
                # Sayısal değerleri formatla
                if isinstance(val, float):
                    val = f"{val:.2f}"
                values.append(val)
            self.tree.insert('', 'end', values=values)
    
    def export_to_excel(self):
        """Excel'e aktar"""
        if not self.current_report_data:
            messagebox.showwarning("Uyarı", "Önce bir rapor oluşturun.")
            return
        
        # Dosya adı sor
        report_type_names = {
            'malzeme_giris': 'Malzeme_Giris',
            'malzeme_cikis': 'Malzeme_Cikis',
            'urun_bazli': 'Urun_Bazli',
            'gunluk_ozet': 'Gunluk_Ozet'
        }
        
        report_name = report_type_names.get(self.report_type_var.get(), 'Rapor')
        default_filename = f"{report_name}_{date.today().strftime('%Y%m%d')}.xlsx"
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=default_filename
        )
        
        if not filepath:
            return
        
        try:
            # Excel workbook oluştur
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rapor"
            
            # Başlık satırı
            columns = list(self.tree["columns"])
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Veri satırları
            for item in self.tree.get_children():
                values = self.tree.item(item)['values']
                ws.append(values)
            
            # Kolon genişliklerini ayarla
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Dosyayı kaydet
            wb.save(filepath)
            
            messagebox.showinfo("Başarılı", f"Rapor Excel'e aktarıldı:\n{filepath}")
            
        except Exception as e:
            messagebox.showerror("Hata", f"Excel'e aktarma hatası:\n{e}")
